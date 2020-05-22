# All your utils are here 
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

def func_export(request, datas):
    ts = str(datetime.now().date())
    filename = "export_" + ts + ".xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Edupay'

    # Define the titles for columns
    columns = [
        "Headder 1",
        "Headder 2",
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for data in datas:

        # change date format
        # coldate = data.ts.date()
        # coldate = coldate.strftime('%d/%m/%Y')

        row = [

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(filename=filename)

    response = HttpResponse(content=save_virtual_workbook(workbook), content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename='+filename
    return response
